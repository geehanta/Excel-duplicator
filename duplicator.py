import openpyxl
import tkinter as tk
from tkinter import filedialog

def duplicate_cells():
    input_file = input_file_entry.get()
    output_file = output_file_entry.get()
    num_duplicates = int(duplicate_entry.get())

    try:
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active

        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active

        for row in ws.iter_rows():
            for cell in row:
                for i in range(num_duplicates):
                    new_cell = new_ws.cell(row=cell.row, column=cell.column + i + 1)
                    new_cell.value = cell.value

        new_wb.save(output_file)
        status_label.config(text="File saved successfully!")
    except Exception as e:
        status_label.config(text=f"Error: {str(e)}")

def browse_input_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    input_file_entry.delete(0, tk.END)
    input_file_entry.insert(0, file_path)

def browse_output_file():
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    output_file_entry.delete(0, tk.END)
    output_file_entry.insert(0, file_path)

# Create GUI window
root = tk.Tk()
root.title("Excel Cell Duplicator")

# Labels and entry fields
tk.Label(root, text="Input Excel File:").grid(row=0, column=0, padx=10, pady=5)
input_file_entry = tk.Entry(root)
input_file_entry.grid(row=0, column=1, padx=10, pady=5)
tk.Button(root, text="Browse", command=browse_input_file).grid(row=0, column=2, padx=5, pady=5)

tk.Label(root, text="Output Excel File:").grid(row=1, column=0, padx=10, pady=5)
output_file_entry = tk.Entry(root)
output_file_entry.grid(row=1, column=1, padx=10, pady=5)
tk.Button(root, text="Browse", command=browse_output_file).grid(row=1, column=2, padx=5, pady=5)

tk.Label(root, text="Number of Duplicates:").grid(row=2, column=0, padx=10, pady=5)
duplicate_entry = tk.Entry(root)
duplicate_entry.grid(row=2, column=1, padx=10, pady=5)

tk.Button(root, text="Duplicate Cells", command=duplicate_cells).grid(row=3, columnspan=3, padx=10, pady=10)

status_label = tk.Label(root, text="", fg="green")
status_label.grid(row=4, columnspan=3, padx=10, pady=5)

root.mainloop()
